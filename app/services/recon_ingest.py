"""app/services/recon_ingest.py

Excel 3-sheet 解析 + 落库 + version supersede。

Sheet 布局：
  Differences     — 差异明细（ReconResultDiff）
  WeDAP Source    — WeDAP 侧源数据（ReconResultSourceWedap）
  Bank Source     — 银行侧源数据（ReconResultSourceBank）

常量：
  PARSER_VERSION  = "1"
  SCHEMA_VERSION  = "wedap-recon-3sheet-v1"
"""

from __future__ import annotations

import logging
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy import update
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.models.recon import (
    ReconResultDiff,
    ReconResultSourceBank,
    ReconResultSourceWedap,
    ReconResultTask,
)


class DataQualityError(Exception):
    """非空金额字段无法解析为 Decimal（脏数据，应触发 FAILED 置位）。"""

    def __init__(self, column: str, value: Any) -> None:
        super().__init__(f"column={column!r} value={value!r} cannot be parsed as Decimal")
        self.column = column
        self.value = value


logger = logging.getLogger(__name__)

PARSER_VERSION = "1"
SCHEMA_VERSION = "wedap-recon-3sheet-v1"

_EXPECTED: dict[str, list[str]] = {
    "Differences": [
        "Type",
        "WeDAP Biz Seq No",
        "Bank Seq No",
        "WeDAP Amount",
        "Bank Amount",
        "Diff Amount",
        "WeDAP Status",
        "Bank Status",
    ],
    "WeDAP Source": [
        "Biz Type",
        "Biz Seq No",
        "Bank Biz Seq No",
        "Amount",
        "Currency",
        "Payer",
        "Payee",
        "Status",
        "Error",
    ],
    "Bank Source": [
        "Bank Seq No",
        "Txn Date",
        "Amount",
        "Currency",
        "Payer",
        "Payee",
        "Status",
        "File Name",
        "Line No",
    ],
}


class ColumnDrift(Exception):
    """Sheet 列头与期望不符，或必要 Sheet 缺失。"""


def _dec(v: Any, column: str | None = None) -> Decimal | None:
    """将单元格值转为 Decimal；空值返回 None；非空非法值抛 DataQualityError。

    Args:
        v: 单元格原值。
        column: 列名，仅在非空解析失败时填入 DataQualityError。
    """
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except InvalidOperation as exc:
        if column is not None:
            raise DataQualityError(column, v) from exc
        return None


def _str(v: Any) -> str | None:
    """将单元格值转为 str；空值返回 None。"""
    if v is None or v == "":
        return None
    return str(v)


def _int(v: Any) -> int | None:
    """将单元格值转为 int；空值返回 None。"""
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def _collect_rows(ws: Any) -> list[tuple[Any, ...]]:
    """跳过首行（列头）+ 跳过全空行，返回数据行元组列表。"""
    result = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        if all(c is None or c == "" for c in row):
            continue
        result.append(row)
    return result


def _check_header(wb_sheets: dict[str, Any], column_check: dict[str, Any]) -> None:
    """校验三个 sheet 列头；有问题时更新 column_check 并抛 ColumnDrift。"""
    for sheet_name, expected_cols in _EXPECTED.items():
        if sheet_name not in wb_sheets:
            column_check["drift"] = f"sheet '{sheet_name}' missing"
            raise ColumnDrift(f"Sheet '{sheet_name}' not found in workbook")
        ws = wb_sheets[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        actual = [str(c) if c is not None else "" for c in header_row]
        if actual != expected_cols:
            column_check["drift"] = f"sheet '{sheet_name}': expected {expected_cols}, got {actual}"
            raise ColumnDrift(
                f"Column drift in '{sheet_name}': expected {expected_cols}, got {actual}"
            )


async def parse_and_land(
    factory: async_sessionmaker[AsyncSession],
    *,
    task_id: int,
    xlsx_path: str,
) -> None:
    """读取 xlsx，解析三个 sheet，落库，并 supersede 同 task_no 的低版本。

    成功：task → PARSED，三子表批量 insert，低版本（status != SUPERSEDED）→ SUPERSEDED。
    失败（ColumnDrift）：task → FAILED，column_check 记录 drift 信息，re-raise。
    失败（DataQualityError）：task → FAILED，column_check 记录 data_error，re-raise。
    """
    # ── 1. 读取 task 基础信息（入口预检查，无锁；事务内终检为准）──────────
    async with factory() as session:
        task = await session.get(ReconResultTask, task_id)
        if task is None:
            raise ValueError(f"ReconResultTask id={task_id} not found")
        # 幂等防御入口预检查：非 DOWNLOADED 状态直接跳过
        if task.status not in ("DOWNLOADED",):
            logger.warning(
                "parse_and_land called on task_id=%s with status=%s; skipping (idempotent)",
                task_id,
                task.status,
            )
            return
        tenant_id: str = task.tenant_id
        task_no: str = task.task_no
        current_version: int = task.version
        expected_diff_count: int = task.diff_count

    # ── 2. 解析 xlsx（read_only + data_only，减少内存占用）────────────────
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    column_check: dict[str, Any] = {}

    try:
        wb_sheets: dict[str, Any] = {ws.title: ws for ws in wb.worksheets}

        try:
            _check_header(wb_sheets, column_check)
        except ColumnDrift:
            async with factory() as session:
                async with session.begin():
                    t = await session.get(ReconResultTask, task_id)
                    if t is not None:  # pragma: no branch
                        t.status = "FAILED"
                        t.column_check = column_check
            raise

        # ── 3. 收集数据行 ──────────────────────────────────────────────────────
        diff_rows = _collect_rows(wb_sheets["Differences"])
        wedap_rows = _collect_rows(wb_sheets["WeDAP Source"])
        bank_rows = _collect_rows(wb_sheets["Bank Source"])
    finally:
        wb.close()

    # 标记三个 sheet 列头均 ok
    column_check["Differences"] = "ok"
    column_check["WeDAP Source"] = "ok"
    column_check["Bank Source"] = "ok"

    # count_mismatch 告警（不阻断解析）
    actual_diff_count = len(diff_rows)
    if actual_diff_count != expected_diff_count:
        column_check["count_mismatch"] = (
            f"expected={expected_diff_count} actual={actual_diff_count}"
        )

    # ── 4. 单事务：批量 insert + task PARSED + supersede 低版本 ───────────
    # 脏金额（DataQualityError）置 FAILED 必须在主事务回滚、FOR UPDATE 行锁释放之后再做：
    # 主事务用 with_for_update 锁住 task 行，若在主事务仍开着时用另一连接 UPDATE 同一 task 行
    # 置 FAILED，两连接互锁 → MySQL 下 innodb 锁等待超时（自死锁）。SQLite 忽略 FOR UPDATE
    # 故单测发现不了。因此把 FAILED 留痕移到 except（主事务已退出、锁已释放）后执行。
    try:
        async with factory() as session:
            async with session.begin():
                # 4-pre. 事务内终检：with_for_update 重读，仅 DOWNLOADED 可继续
                t_check = await session.get(ReconResultTask, task_id, with_for_update=True)
                if t_check is None or t_check.status not in ("DOWNLOADED",):
                    logger.warning(
                        "parse_and_land txn-guard: task_id=%s status=%s; skipping",
                        task_id,
                        t_check.status if t_check is not None else "GONE",
                    )
                    return

                # 4a. 差异明细
                for _row_idx, row in enumerate(diff_rows):
                    (
                        diff_type,
                        wedap_biz_seq_no,
                        bank_seq_no,
                        wedap_amount,
                        bank_amount,
                        diff_amount,
                        wedap_status,
                        bank_status,
                    ) = row[:8]
                    session.add(
                        ReconResultDiff(
                            tenant_id=tenant_id,
                            task_id=task_id,
                            diff_type=_str(diff_type) or "",
                            wedap_biz_seq_no=_str(wedap_biz_seq_no),
                            bank_seq_no=_str(bank_seq_no),
                            wedap_amount=_dec(wedap_amount, "wedap_amount"),
                            bank_amount=_dec(bank_amount, "bank_amount"),
                            diff_amount=_dec(diff_amount, "diff_amount"),
                            wedap_status=_str(wedap_status),
                            bank_status=_str(bank_status),
                        )
                    )

                # 4b. WeDAP 源数据
                for _row_idx, row in enumerate(wedap_rows):
                    (
                        biz_type,
                        biz_seq_no,
                        bank_biz_seq_no,
                        amount,
                        currency,
                        payer,
                        payee,
                        status,
                        error,
                    ) = row[:9]
                    session.add(
                        ReconResultSourceWedap(
                            tenant_id=tenant_id,
                            task_id=task_id,
                            biz_type=_str(biz_type),
                            biz_seq_no=_str(biz_seq_no) or "",
                            bank_biz_seq_no=_str(bank_biz_seq_no),
                            amount=_dec(amount, "amount"),
                            currency=_str(currency),
                            payer_account=_str(payer),
                            payee_account=_str(payee),
                            status=_str(status),
                            error_msg=_str(error),
                        )
                    )

                # 4c. 银行源数据
                for _row_idx, row in enumerate(bank_rows):
                    (
                        bank_seq_no,
                        txn_date,
                        amount,
                        currency,
                        payer,
                        payee,
                        status,
                        file_name,
                        line_no,
                    ) = row[:9]
                    session.add(
                        ReconResultSourceBank(
                            tenant_id=tenant_id,
                            task_id=task_id,
                            bank_seq_no=_str(bank_seq_no) or "",
                            txn_date=_str(txn_date),
                            amount=_dec(amount, "amount"),
                            currency=_str(currency),
                            payer_account=_str(payer),
                            payee_account=_str(payee),
                            status=_str(status),
                            file_name=_str(file_name),
                            line_no=_int(line_no),
                        )
                    )

                # 4d. 更新 task 为 PARSED
                t_check.status = "PARSED"
                t_check.parser_version = PARSER_VERSION
                t_check.schema_version = SCHEMA_VERSION
                t_check.column_check = column_check
                t_check.archive_path = xlsx_path

                # 4e. 同 tenant_id + task_no 低版本（status != SUPERSEDED）→ SUPERSEDED
                # P1-1 修复：加 tenant_id 条件，防止跨租户污染
                await session.execute(
                    update(ReconResultTask)
                    .where(
                        ReconResultTask.tenant_id == tenant_id,
                        ReconResultTask.task_no == task_no,
                        ReconResultTask.version < current_version,
                        ReconResultTask.status != "SUPERSEDED",
                    )
                    .values(status="SUPERSEDED")
                )
    except DataQualityError as dqe:
        # 脏金额不静默：主事务已回滚（三表 0 行、FOR UPDATE 锁已释放），独立事务置 FAILED 留痕
        dq_payload: dict[str, Any] = {"data_error": f"column={dqe.column!r} value={dqe.value!r}"}
        async with factory() as _sess:
            async with _sess.begin():
                _t = await _sess.get(ReconResultTask, task_id)
                if _t is not None:  # pragma: no branch
                    _t.status = "FAILED"
                    _t.column_check = dq_payload
        raise
