"""集成测试（需 docker）：在真实 MySQL 8.0 容器中验证 recon parse_and_land 的脏金额路径
不会因「主事务持 task 行 FOR UPDATE 锁 + 另一连接 UPDATE 同行置 FAILED」而自死锁。

背景（team-review iteration-1 finding REC-DEADLOCK）：
parse_and_land 在 `with_for_update` 锁住 task 行的主事务内解析三表；旧实现遇脏金额时，
在 except 中开**另一连接**对同一 task 行 UPDATE status='FAILED'。主事务此刻仍开着、持
有该行排他锁 → 两连接互锁 → MySQL innodb 锁等待超时（自死锁）。SQLite 忽略 FOR UPDATE，
单测发现不了。修复：把 FAILED 留痕移到主事务回滚（锁释放）之后执行。

本测试把 innodb_lock_wait_timeout 调到 5s：
- 旧（有 bug）实现：第二连接 UPDATE 卡锁 5s → OperationalError（lock wait timeout），
  且 task 不会被置 FAILED。
- 新（修复）实现：抛 DataQualityError（秒级），task 置 FAILED，三子表 0 行。

运行方式：
    .venv/bin/pytest -m integration --no-cov -q tests/integration/test_recon_ingest_mysql.py
本机无 docker / asyncmy 时跳过。
"""

from __future__ import annotations

import pytest
import sqlalchemy as sa
from openpyxl import Workbook

pytestmark = pytest.mark.integration

DIFF_HEADER = [
    "Type",
    "WeDAP Biz Seq No",
    "Bank Seq No",
    "WeDAP Amount",
    "Bank Amount",
    "Diff Amount",
    "WeDAP Status",
    "Bank Status",
]
WEDAP_HEADER = [
    "Biz Type",
    "Biz Seq No",
    "Bank Biz Seq No",
    "Amount",
    "Currency",
    "Payer",
    "Payee",
    "Status",
    "Error",
]
BANK_HEADER = [
    "Bank Seq No",
    "Txn Date",
    "Amount",
    "Currency",
    "Payer",
    "Payee",
    "Status",
    "File Name",
    "Line No",
]


def _make_dirty_xlsx(path: str) -> None:
    """三 sheet 列头合法，但 WeDAP Source 的 Amount 列含无法解析为 Decimal 的脏值。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Differences"
    ws.append(DIFF_HEADER)
    w2 = wb.create_sheet("WeDAP Source")
    w2.append(WEDAP_HEADER)
    # Amount 列 = "not-a-number" → _dec(..., "amount") 抛 DataQualityError
    w2.append(["DSB", "BIZ-001", "BANK-001", "not-a-number", "USD", "p", "q", "OK", None])
    w3 = wb.create_sheet("Bank Source")
    w3.append(BANK_HEADER)
    wb.save(path)


@pytest.fixture(scope="module")
def mysql_async_url() -> str:
    """启动 MySQL 8.0 testcontainer，建 recon 表，把 innodb_lock_wait_timeout 调到 5s，
    返回 root asyncmy URL（test 用户在 WSL2 docker bridge 下 asyncmy 连接会 Access denied）。"""
    try:
        from testcontainers.mysql import MySqlContainer  # type: ignore[import-untyped]
    except ImportError:
        pytest.skip("testcontainers 未安装")
    try:
        import asyncmy  # noqa: F401
    except ImportError:
        pytest.skip("asyncmy 未安装")

    # import 确保 recon 模型注册到 Base.metadata
    import app.models.recon  # noqa: F401
    from app.models.base import Base

    try:
        with MySqlContainer("mysql:8.0") as mysql:
            host = mysql.get_container_host_ip()
            port = mysql.get_exposed_port(3306)
            db = mysql.dbname
            root_sync = f"mysql+pymysql://root:test@{host}:{port}/{db}"

            sync_engine = sa.create_engine(root_sync, pool_pre_ping=True)
            # 锁等待超时调小，确保（旧实现下）自死锁能快速暴露而非 50s 阻塞
            with sync_engine.connect() as conn:
                conn.execute(sa.text("SET GLOBAL innodb_lock_wait_timeout = 5"))
                conn.commit()
            Base.metadata.create_all(sync_engine)
            sync_engine.dispose()

            yield f"mysql+asyncmy://root:test@{host}:{port}/{db}"
    except Exception as exc:  # pragma: no cover
        pytest.skip(f"docker/testcontainer 不可用: {exc}")


@pytest.mark.asyncio
async def test_dirty_amount_no_self_deadlock_marks_failed(mysql_async_url: str, tmp_path) -> None:
    """脏金额：parse_and_land 抛 DataQualityError（非锁等待超时），task 置 FAILED，三表 0 行。

    旧实现会在此触发 FOR UPDATE 自死锁（OperationalError lock wait timeout）且不置 FAILED。
    """
    from app.core.db import build_engine, build_session_factory
    from app.models.recon import (
        ReconResultDiff,
        ReconResultSourceBank,
        ReconResultSourceWedap,
        ReconResultTask,
    )
    from app.services.recon_ingest import DataQualityError, parse_and_land

    engine = build_engine(mysql_async_url)
    factory = build_session_factory(engine)
    try:
        # 插入一个 DOWNLOADED task
        async with factory() as session:
            async with session.begin():
                task = ReconResultTask(
                    tenant_id="WBTHK01",
                    task_no="RCN-DEADLOCK-001",
                    version=1,
                    recon_date="20260615",
                    s3_bucket="b",
                    s3_key="k",
                    file_md5="m",
                    diff_count=0,
                    status="DOWNLOADED",
                    request_id="REQ-DEADLOCK-001",
                )
                session.add(task)
            await session.refresh(task)
            task_id = task.id

        xlsx = tmp_path / "dirty.xlsx"
        _make_dirty_xlsx(str(xlsx))

        # 关键断言：抛 DataQualityError（修复后），而不是 lock wait timeout（OperationalError）
        with pytest.raises(DataQualityError):
            await parse_and_land(factory, task_id=task_id, xlsx_path=str(xlsx))

        # task 置 FAILED + data_error 留痕；三子表 0 行（主事务回滚）
        async with factory() as session:
            t = await session.get(ReconResultTask, task_id)
            assert t is not None
            assert t.status == "FAILED"
            assert t.column_check is not None and "data_error" in t.column_check

            diff_n = await session.scalar(sa.select(sa.func.count()).select_from(ReconResultDiff))
            wedap_n = await session.scalar(
                sa.select(sa.func.count()).select_from(ReconResultSourceWedap)
            )
            bank_n = await session.scalar(
                sa.select(sa.func.count()).select_from(ReconResultSourceBank)
            )
            assert (diff_n, wedap_n, bank_n) == (0, 0, 0)
    finally:
        await engine.dispose()
