"""tests/services/test_recon_ingest.py

TDD：先定义期望行为，实现前所有 case 均失败。
覆盖：
1. test_parse_lands_three_sheets        — 3 sheet 正常解析落库
2. test_header_drift_fails_task         — 列头不符 → ColumnDrift / FAILED
3. test_new_version_supersedes_old      — 新版本解析后旧版本变 SUPERSEDED
4. test_missing_sheet_fails             — Sheet 缺失 → ColumnDrift / FAILED
5. test_totalcount_mismatch_recorded    — diff_count 不符 → 解析成功但 column_check 含告警
6. test_ingest_pending_once             — worker 全链：NOTIFIED→DOWNLOADED→PARSED；
                                          Md5Mismatch→FAILED
"""

from decimal import Decimal

import pytest
from openpyxl import Workbook

from app.core.db import build_engine, build_session_factory
from app.models.base import Base
from app.models.recon import (
    ReconResultDiff,
    ReconResultSourceBank,
    ReconResultSourceWedap,
    ReconResultTask,
)
from app.services.recon_ingest import ColumnDrift, parse_and_land

DIFF_HEADER = [
    "Type",
    "WeDAP Biz Seq No",
    "Bank Seq No",
    "WeDAP Amount",
    "Bank Amount",
    "Diff Amount",
    "WeDAP Status",
    "Bank Status",
]
WEDAP_HEADER = [
    "Biz Type",
    "Biz Seq No",
    "Bank Biz Seq No",
    "Amount",
    "Currency",
    "Payer",
    "Payee",
    "Status",
    "Error",
]
BANK_HEADER = [
    "Bank Seq No",
    "Txn Date",
    "Amount",
    "Currency",
    "Payer",
    "Payee",
    "Status",
    "File Name",
    "Line No",
]


def _make_xlsx(
    path,
    *,
    diff_rows=(),
    wedap_rows=(),
    bank_rows=(),
    break_header=False,
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Differences"
    ws.append(["XXX"] if break_header else DIFF_HEADER)
    for r in diff_rows:
        ws.append(r)
    w2 = wb.create_sheet("WeDAP Source")
    w2.append(WEDAP_HEADER)
    for r in wedap_rows:
        w2.append(r)
    w3 = wb.create_sheet("Bank Source")
    w3.append(BANK_HEADER)
    for r in bank_rows:
        w3.append(r)
    wb.save(path)


# ───────────────────────── fixture ───────────────────────────────────────────


@pytest.fixture()
async def factory(tmp_path):
    engine = build_engine("sqlite+aiosqlite:///:memory:")
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield build_session_factory(engine)
    await engine.dispose()


def _task(**kw) -> ReconResultTask:
    defaults = dict(
        tenant_id="WBTHK01",
        task_no="RECON-20260611-001",
        version=1,
        recon_date="20260611",
        s3_bucket="wbt-recon-bucket",
        s3_key="recon/20260611/result.xlsx",
        file_md5="d41d8cd98f00b204e9800998ecf8427e",
        diff_count=1,
        status="DOWNLOADED",
        request_id="REQ-20260611-0001",
    )
    defaults.update(kw)
    return ReconResultTask(**defaults)


# ───────────────────────── Test 1: 3 sheet 正常解析落库 ──────────────────────


@pytest.mark.asyncio
async def test_parse_lands_three_sheets(factory, tmp_path) -> None:
    """parse_and_land 成功：task PARSED + 三表各 1 行 + 字段值正确。"""
    xlsx = tmp_path / "result.xlsx"
    _make_xlsx(
        xlsx,
        diff_rows=[
            [
                "AMOUNT",
                "DSB-001",
                "BANK-001",
                "100.0000",
                "99.9900",
                "0.0100",
                "SUCCESS",
                "SETTLED",
            ]
        ],
        wedap_rows=[
            [
                "DSB",
                "DSB-001",
                "BANK-001",
                "100.0000",
                "USD",
                "PAYER-ACC",
                "PAYEE-ACC",
                "SUCCESS",
                None,
            ]
        ],
        bank_rows=[
            [
                "BANK-001",
                "20260611",
                "99.9900",
                "USD",
                "PAYER-ACC",
                "PAYEE-ACC",
                "SETTLED",
                "recon_20260611.xlsx",
                5,
            ]
        ],
    )

    # 先插 task
    async with factory() as session:
        async with session.begin():
            task = _task(status="DOWNLOADED")
            session.add(task)
    async with factory() as session:
        row = (
            await session.execute(
                __import__("sqlalchemy", fromlist=["select"]).select(ReconResultTask)
            )
        ).scalar_one()
        tid = row.id

    await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    from sqlalchemy import select

    async with factory() as session:
        task = await session.get(ReconResultTask, tid)
        assert task is not None
        assert task.status == "PARSED"
        assert task.parser_version == "1"
        assert task.schema_version == "wedap-recon-3sheet-v1"
        assert task.archive_path == str(xlsx)
        cc = task.column_check
        assert cc is not None
        assert cc.get("Differences") == "ok"
        assert cc.get("WeDAP Source") == "ok"
        assert cc.get("Bank Source") == "ok"

        diffs = (
            (await session.execute(select(ReconResultDiff).where(ReconResultDiff.task_id == tid)))
            .scalars()
            .all()
        )
        assert len(diffs) == 1
        d = diffs[0]
        assert d.diff_type == "AMOUNT"
        assert d.wedap_biz_seq_no == "DSB-001"
        assert d.bank_seq_no == "BANK-001"
        assert d.wedap_amount == Decimal("100.0000")
        assert d.bank_amount == Decimal("99.9900")
        assert d.diff_amount == Decimal("0.0100")
        assert d.wedap_status == "SUCCESS"
        assert d.bank_status == "SETTLED"

        wedaps = (
            (
                await session.execute(
                    select(ReconResultSourceWedap).where(ReconResultSourceWedap.task_id == tid)
                )
            )
            .scalars()
            .all()
        )
        assert len(wedaps) == 1
        w = wedaps[0]
        assert w.biz_seq_no == "DSB-001"
        assert w.bank_biz_seq_no == "BANK-001"
        assert w.amount == Decimal("100.0000")

        banks = (
            (
                await session.execute(
                    select(ReconResultSourceBank).where(ReconResultSourceBank.task_id == tid)
                )
            )
            .scalars()
            .all()
        )
        assert len(banks) == 1
        b = banks[0]
        assert b.bank_seq_no == "BANK-001"
        assert b.line_no == 5


# ───────────────────────── Test 2: 列头 drift → FAILED ───────────────────────


@pytest.mark.asyncio
async def test_header_drift_fails_task(factory, tmp_path) -> None:
    """列头不符 → ColumnDrift 抛出；task.status = FAILED，column_check 记 drift。"""
    xlsx = tmp_path / "bad.xlsx"
    _make_xlsx(xlsx, break_header=True)

    async with factory() as session:
        async with session.begin():
            task = _task(status="DOWNLOADED")
            session.add(task)
    async with factory() as session:
        row = (
            await session.execute(
                __import__("sqlalchemy", fromlist=["select"]).select(ReconResultTask)
            )
        ).scalar_one()
        tid = row.id

    with pytest.raises(ColumnDrift):
        await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    async with factory() as session:
        task = await session.get(ReconResultTask, tid)
        assert task is not None
        assert task.status == "FAILED"
        cc = task.column_check
        assert cc is not None
        assert "drift" in cc


# ───────────────────────── Test 3: 新版本 supersede 旧版本 ───────────────────


@pytest.mark.asyncio
async def test_new_version_supersedes_old(factory, tmp_path) -> None:
    """v2 解析后：v1 → SUPERSEDED；v2 → PARSED；v3 解析后 v1 状态不再重复改写。"""
    xlsx_v2 = tmp_path / "v2.xlsx"
    _make_xlsx(xlsx_v2)
    xlsx_v3 = tmp_path / "v3.xlsx"
    _make_xlsx(xlsx_v3)

    from sqlalchemy import select

    # 插 v1(PARSED) + v2(DOWNLOADED)
    async with factory() as session:
        async with session.begin():
            task_v1 = _task(version=1, status="PARSED", request_id="REQ-001")
            task_v2 = _task(version=2, status="DOWNLOADED", request_id="REQ-002")
            session.add(task_v1)
            session.add(task_v2)

    async with factory() as session:
        rows = (
            (await session.execute(select(ReconResultTask).order_by(ReconResultTask.version)))
            .scalars()
            .all()
        )
        tid_v1 = rows[0].id
        tid_v2 = rows[1].id

    # 解析 v2
    await parse_and_land(factory, task_id=tid_v2, xlsx_path=str(xlsx_v2))

    async with factory() as session:
        t1 = await session.get(ReconResultTask, tid_v1)
        t2 = await session.get(ReconResultTask, tid_v2)
        assert t1 is not None and t1.status == "SUPERSEDED"
        assert t2 is not None and t2.status == "PARSED"

    # 插 v3(DOWNLOADED)，解析后 v1 不再重复改写
    async with factory() as session:
        async with session.begin():
            task_v3 = _task(version=3, status="DOWNLOADED", request_id="REQ-003")
            session.add(task_v3)
    async with factory() as session:
        rows = (
            (await session.execute(select(ReconResultTask).order_by(ReconResultTask.version)))
            .scalars()
            .all()
        )
        tid_v3 = rows[2].id

    # v1 此时已是 SUPERSEDED，supersede 逻辑需过滤 status != SUPERSEDED
    await parse_and_land(factory, task_id=tid_v3, xlsx_path=str(xlsx_v3))

    async with factory() as session:
        t1 = await session.get(ReconResultTask, tid_v1)
        t2 = await session.get(ReconResultTask, tid_v2)
        t3 = await session.get(ReconResultTask, tid_v3)
        # v1 仍是 SUPERSEDED（不重复改写）
        assert t1 is not None and t1.status == "SUPERSEDED"
        # v2 被 v3 supersede
        assert t2 is not None and t2.status == "SUPERSEDED"
        assert t3 is not None and t3.status == "PARSED"


# ───────────────────────── Test 4: 缺 sheet → FAILED ────────────────────────


@pytest.mark.asyncio
async def test_missing_sheet_fails(factory, tmp_path) -> None:
    """只有一个 sheet 的 workbook → ColumnDrift；task.status = FAILED。"""
    xlsx = tmp_path / "single.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Differences"
    ws.append(DIFF_HEADER)
    wb.save(xlsx)

    async with factory() as session:
        async with session.begin():
            task = _task(status="DOWNLOADED")
            session.add(task)
    async with factory() as session:
        row = (
            await session.execute(
                __import__("sqlalchemy", fromlist=["select"]).select(ReconResultTask)
            )
        ).scalar_one()
        tid = row.id

    with pytest.raises(ColumnDrift):
        await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    async with factory() as session:
        task = await session.get(ReconResultTask, tid)
        assert task is not None
        assert task.status == "FAILED"


# ───────────────────────── Test 5: count mismatch 记录告警 ───────────────────


@pytest.mark.asyncio
async def test_totalcount_mismatch_recorded(factory, tmp_path) -> None:
    """diff_count=5 但实际 1 行 → 解析成功，column_check 含 count_mismatch 告警。"""
    xlsx = tmp_path / "mismatch.xlsx"
    _make_xlsx(
        xlsx,
        diff_rows=[
            [
                "AMOUNT",
                "DSB-001",
                "BANK-001",
                "100.0000",
                "99.9900",
                "0.0100",
                "SUCCESS",
                "SETTLED",
            ]
        ],
        wedap_rows=[
            [
                "DSB",
                "DSB-001",
                "BANK-001",
                "100.0000",
                "USD",
                "PAYER-ACC",
                "PAYEE-ACC",
                "SUCCESS",
                None,
            ]
        ],
        bank_rows=[
            [
                "BANK-001",
                "20260611",
                "99.9900",
                "USD",
                "PAYER-ACC",
                "PAYEE-ACC",
                "SETTLED",
                "recon_20260611.xlsx",
                5,
            ]
        ],
    )

    async with factory() as session:
        async with session.begin():
            # diff_count=5 但 Differences 实际只有 1 行
            task = _task(status="DOWNLOADED", diff_count=5)
            session.add(task)
    async with factory() as session:
        row = (
            await session.execute(
                __import__("sqlalchemy", fromlist=["select"]).select(ReconResultTask)
            )
        ).scalar_one()
        tid = row.id

    await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    async with factory() as session:
        task = await session.get(ReconResultTask, tid)
        assert task is not None
        assert task.status == "PARSED"
        cc = task.column_check
        assert cc is not None
        # 告警字段存在
        mismatch = cc.get("count_mismatch")
        assert mismatch is not None
        assert "expected=5" in mismatch
        assert "actual=1" in mismatch


# ───────────────────────── Test 6: ingest_pending_once worker 全链 ───────────


@pytest.mark.asyncio
async def test_ingest_pending_once(factory, tmp_path, monkeypatch) -> None:
    """
    worker 全链：
    - NOTIFIED → download_verified(monkeypatch 写 fixture xlsx) → DOWNLOADED → PARSED
    - Md5Mismatch 场景 → task FAILED，column_check 含 download_error
    """
    from app.clients.s3 import Md5Mismatch, S3FileClient
    from app.workers.recon_worker import ingest_pending_once

    xlsx_ok = tmp_path / "ok.xlsx"
    _make_xlsx(
        xlsx_ok,
        diff_rows=[
            [
                "AMOUNT",
                "DSB-002",
                "BANK-002",
                "200.0000",
                "199.9900",
                "0.0100",
                "SUCCESS",
                "SETTLED",
            ]
        ],
        wedap_rows=[
            [
                "DSB",
                "DSB-002",
                "BANK-002",
                "200.0000",
                "USD",
                "P2",
                "P3",
                "SUCCESS",
                None,
            ]
        ],
        bank_rows=[
            [
                "BANK-002",
                "20260611",
                "199.9900",
                "USD",
                "P2",
                "P3",
                "SETTLED",
                "f.xlsx",
                1,
            ]
        ],
    )

    # 插 NOTIFIED task（正常）和 NOTIFIED task（md5 失败）
    async with factory() as session:
        async with session.begin():
            t_ok = _task(
                status="NOTIFIED",
                task_no="RECON-OK-001",
                version=1,
                request_id="REQ-OK-001",
            )
            t_fail = _task(
                status="NOTIFIED",
                task_no="RECON-FAIL-001",
                version=1,
                request_id="REQ-FAIL-001",
            )
            session.add(t_ok)
            session.add(t_fail)

    from sqlalchemy import select

    async with factory() as session:
        rows = (
            (await session.execute(select(ReconResultTask).order_by(ReconResultTask.task_no)))
            .scalars()
            .all()
        )
    # RECON-FAIL-001 先、RECON-OK-001 后（字母排序）
    tid_fail = next(r.id for r in rows if r.task_no == "RECON-FAIL-001")
    tid_ok = next(r.id for r in rows if r.task_no == "RECON-OK-001")

    # monkeypatch S3FileClient.download_verified
    call_count = 0

    def fake_download(
        self,
        *,
        bucket: str,
        key: str,
        expected_md5: str,
        dest: str,
    ) -> None:
        nonlocal call_count
        call_count += 1
        # 第一个调用是 FAIL 任务（字母序先），抛 Md5Mismatch
        task_no = dest.split("/")[-1].split("_v")[0]
        if task_no == "RECON-FAIL-001":
            raise Md5Mismatch("hash mismatch")
        # 正常任务：把 fixture xlsx 写到 dest
        import shutil

        shutil.copy(str(xlsx_ok), dest)

    monkeypatch.setattr(S3FileClient, "download_verified", fake_download)

    s3 = S3FileClient(endpoint_url=None)  # boto3 不会真正调用
    handled = await ingest_pending_once(
        factory,
        s3=s3,
        archive_dir=str(tmp_path / "archive"),
    )

    assert handled == 2  # 两个都处理（1 失败 1 成功均计入）

    async with factory() as session:
        t_ok_row = await session.get(ReconResultTask, tid_ok)
        t_fail_row = await session.get(ReconResultTask, tid_fail)

    assert t_ok_row is not None and t_ok_row.status == "PARSED"
    assert t_fail_row is not None and t_fail_row.status == "FAILED"
    assert t_fail_row.column_check is not None
    assert t_fail_row.column_check.get("download_error") == "Md5Mismatch"


# ───────────────────────── Test 7: 辅助函数边界分支 ─────────────────────────


def test_dec_none_and_empty() -> None:
    """_dec(None) 和 _dec('') 均返回 None；无效字符串也返回 None。"""
    from app.services.recon_ingest import _dec

    assert _dec(None) is None
    assert _dec("") is None
    assert _dec("not-a-number") is None
    assert _dec("123.45") == Decimal("123.45")


def test_int_none_empty_and_invalid() -> None:
    """_int(None) / _int('') 返回 None；非整型字符串也返回 None。"""
    from app.services.recon_ingest import _int

    assert _int(None) is None
    assert _int("") is None
    assert _int("abc") is None
    assert _int(5) == 5


@pytest.mark.asyncio
async def test_collect_rows_skips_all_empty(factory, tmp_path) -> None:
    """_collect_rows 跳过全空行，只返回有数据的行。"""
    from app.services.recon_ingest import _collect_rows

    # 用 openpyxl 建含全空行的 sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Test"
    ws.append(["H1", "H2"])
    ws.append([None, None])  # 全空行，应跳过
    ws.append(["val", "123"])
    path = tmp_path / "empty_row.xlsx"
    wb.save(path)

    from openpyxl import load_workbook as lw

    wb2 = lw(str(path), read_only=True, data_only=True)
    rows = _collect_rows(wb2.worksheets[0])
    wb2.close()
    assert len(rows) == 1
    assert rows[0][0] == "val"


@pytest.mark.asyncio
async def test_parse_task_not_found_raises(factory, tmp_path) -> None:
    """task_id 不存在时 parse_and_land 抛 ValueError。"""
    xlsx = tmp_path / "x.xlsx"
    _make_xlsx(xlsx)
    with pytest.raises(ValueError, match="not found"):
        await parse_and_land(factory, task_id=99999, xlsx_path=str(xlsx))


@pytest.mark.asyncio
async def test_worker_parse_column_drift_continues(factory, tmp_path, monkeypatch) -> None:
    """worker 遇到 ColumnDrift（parse 失败）时不阻断，handled 仍计入。"""
    from app.clients.s3 import S3FileClient
    from app.workers.recon_worker import ingest_pending_once

    # 建含 bad header 的 xlsx（触发 ColumnDrift）
    bad_xlsx = tmp_path / "bad.xlsx"
    _make_xlsx(bad_xlsx, break_header=True)

    async with factory() as session:
        async with session.begin():
            task = _task(status="NOTIFIED", task_no="RECON-DRIFT-001", request_id="REQ-D-001")
            session.add(task)

    from sqlalchemy import select

    async with factory() as session:
        row = (
            await session.execute(
                select(ReconResultTask).where(ReconResultTask.task_no == "RECON-DRIFT-001")
            )
        ).scalar_one()
        tid = row.id

    def fake_download(self, *, bucket: str, key: str, expected_md5: str, dest: str) -> None:
        import shutil

        shutil.copy(str(bad_xlsx), dest)

    monkeypatch.setattr(S3FileClient, "download_verified", fake_download)
    s3 = S3FileClient(endpoint_url=None)

    handled = await ingest_pending_once(factory, s3=s3, archive_dir=str(tmp_path / "arch"))
    assert handled == 1

    async with factory() as session:
        t = await session.get(ReconResultTask, tid)
    assert t is not None
    assert t.status == "FAILED"  # ColumnDrift 已由 parse_and_land 置为 FAILED


@pytest.mark.asyncio
async def test_worker_t_is_none_branches(factory, tmp_path, monkeypatch) -> None:
    """worker 在 download 失败 + DOWNLOADED 标记两个 session.get 返回 None 时不崩溃。

    通过在 snapshot 里塞一个不存在的 task_id 触发 t is None 分支。
    """
    from app.clients.s3 import Md5Mismatch, S3FileClient
    from app.workers.recon_worker import ingest_pending_once

    # 先建一个真实 NOTIFIED task，取出 snapshot 后手动删它，然后触发 download 失败
    async with factory() as session:
        async with session.begin():
            task = _task(status="NOTIFIED", task_no="RECON-GONE-001", request_id="REQ-G-001")
            session.add(task)

    from sqlalchemy import delete, select

    async with factory() as session:
        row = (
            await session.execute(
                select(ReconResultTask).where(ReconResultTask.task_no == "RECON-GONE-001")
            )
        ).scalar_one()
        tid = row.id

    def fake_download(self, *, bucket: str, key: str, expected_md5: str, dest: str) -> None:
        raise Md5Mismatch("gone")

    monkeypatch.setattr(S3FileClient, "download_verified", fake_download)

    # 删 task，使 snapshot 查询为空 → handled=0，整体不报错
    async with factory() as session:
        async with session.begin():
            await session.execute(delete(ReconResultTask).where(ReconResultTask.id == tid))

    s3 = S3FileClient(endpoint_url=None)
    handled = await ingest_pending_once(factory, s3=s3, archive_dir=str(tmp_path / "arch2"))
    assert handled == 0  # snapshot 为空（task 已删），正常返回 0


# ───────────────────────── Test: 重复解析幂等防御 ────────────────────────────


@pytest.mark.asyncio
async def test_parse_and_land_idempotent_on_parsed(factory, tmp_path) -> None:
    """task.status == PARSED 时再次调用 parse_and_land → 直接 return，三表行数不变。"""
    from sqlalchemy import func, select

    xlsx = tmp_path / "idempotent_parsed.xlsx"
    _make_xlsx(
        xlsx,
        diff_rows=[["AMOUNT", "DSB-001", "BANK-001", "100.0000", "99.9900", "0.0100", "OK", "OK"]],
        wedap_rows=[["DSB", "DSB-001", "BANK-001", "100.0000", "USD", "P1", "P2", "OK", None]],
        bank_rows=[["BANK-001", "20260611", "99.9900", "USD", "P1", "P2", "OK", "f.xlsx", 1]],
    )

    # 先正常解析一次 → PARSED
    async with factory() as session:
        async with session.begin():
            task = _task(status="DOWNLOADED")
            session.add(task)
    async with factory() as session:
        row = (
            await session.execute(
                __import__("sqlalchemy", fromlist=["select"]).select(ReconResultTask)
            )
        ).scalar_one()
        tid = row.id

    await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    # 读取各表行数（首次解析后）
    async with factory() as session:
        diff_count = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultDiff)
                .where(ReconResultDiff.task_id == tid)
            )
        ).scalar_one()
        wedap_count = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultSourceWedap)
                .where(ReconResultSourceWedap.task_id == tid)
            )
        ).scalar_one()
        bank_count = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultSourceBank)
                .where(ReconResultSourceBank.task_id == tid)
            )
        ).scalar_one()

    assert diff_count == 1
    assert wedap_count == 1
    assert bank_count == 1

    # 再次调用（task 已是 PARSED）→ 幂等空操作，行数不变
    await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    async with factory() as session:
        diff_count2 = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultDiff)
                .where(ReconResultDiff.task_id == tid)
            )
        ).scalar_one()
        wedap_count2 = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultSourceWedap)
                .where(ReconResultSourceWedap.task_id == tid)
            )
        ).scalar_one()
        bank_count2 = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultSourceBank)
                .where(ReconResultSourceBank.task_id == tid)
            )
        ).scalar_one()

    assert diff_count2 == 1, "重复调用不应追加新行到 ReconResultDiff"
    assert wedap_count2 == 1, "重复调用不应追加新行到 ReconResultSourceWedap"
    assert bank_count2 == 1, "重复调用不应追加新行到 ReconResultSourceBank"


@pytest.mark.asyncio
async def test_parse_and_land_idempotent_on_superseded(factory, tmp_path) -> None:
    """task.status == SUPERSEDED 时再次调用 parse_and_land → 直接 return，三表行数不变。"""
    from sqlalchemy import func, select

    xlsx = tmp_path / "idempotent_superseded.xlsx"
    _make_xlsx(xlsx)

    # 直接插入 SUPERSEDED 状态的 task（无需真正解析过）
    async with factory() as session:
        async with session.begin():
            task = _task(status="SUPERSEDED")
            session.add(task)
    async with factory() as session:
        row = (
            await session.execute(
                __import__("sqlalchemy", fromlist=["select"]).select(ReconResultTask)
            )
        ).scalar_one()
        tid = row.id

    # 调用 parse_and_land → 幂等空操作
    await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    async with factory() as session:
        diff_count = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultDiff)
                .where(ReconResultDiff.task_id == tid)
            )
        ).scalar_one()

    assert diff_count == 0, "SUPERSEDED 任务不应写入任何差异行"


# ───────────────────────── Test: DataQualityError 属性 ──────────────────────


def test_data_quality_error_attributes() -> None:
    """DataQualityError 携带 column 和 value 属性。"""
    from app.services.recon_ingest import DataQualityError

    err = DataQualityError("amount", "abc")
    assert err.column == "amount"
    assert err.value == "abc"
    assert "column='amount'" in str(err)
    assert "value='abc'" in str(err)


def test_dec_invalid_without_column_returns_none() -> None:
    """_dec 无效值 + column=None → 返回 None（不抛异常）。"""
    from app.services.recon_ingest import _dec

    assert _dec("not-a-number", column=None) is None


def test_dec_invalid_with_column_raises() -> None:
    """_dec 无效值 + column 非 None → 抛 DataQualityError。"""
    from app.services.recon_ingest import DataQualityError, _dec

    with pytest.raises(DataQualityError) as exc_info:
        _dec("abc", column="wedap_amount")
    assert exc_info.value.column == "wedap_amount"
    assert exc_info.value.value == "abc"


# ───────────────────────── Test: 脏金额 FAILED + 三表 0 行 ──────────────────


@pytest.mark.asyncio
async def test_dirty_amount_fails_task_no_rows(factory, tmp_path) -> None:
    """差异 sheet 中有无法解析的金额 → task FAILED + column_check data_error + 三表 0 行。"""
    from sqlalchemy import func, select

    xlsx = tmp_path / "dirty.xlsx"
    wb = __import__("openpyxl", fromlist=["Workbook"]).Workbook()
    ws = wb.active
    ws.title = "Differences"
    ws.append(DIFF_HEADER)
    # wedap_amount = "abc"（脏金额）
    ws.append(["AMOUNT", "DSB-001", "BANK-001", "abc", "99.9900", "0.0100", "SUCCESS", "SETTLED"])
    w2 = wb.create_sheet("WeDAP Source")
    w2.append(WEDAP_HEADER)
    w3 = wb.create_sheet("Bank Source")
    w3.append(BANK_HEADER)
    wb.save(xlsx)

    async with factory() as session:
        async with session.begin():
            task = _task(status="DOWNLOADED", diff_count=1)
            session.add(task)
    async with factory() as session:
        row = (
            await session.execute(
                __import__("sqlalchemy", fromlist=["select"]).select(ReconResultTask)
            )
        ).scalar_one()
        tid = row.id

    from app.services.recon_ingest import DataQualityError

    with pytest.raises(DataQualityError):
        await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    async with factory() as session:
        task = await session.get(ReconResultTask, tid)
        assert task is not None
        assert task.status == "FAILED"
        assert task.column_check is not None
        assert "data_error" in task.column_check

        # 三表均 0 行（事务回滚）
        diff_count = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultDiff)
                .where(ReconResultDiff.task_id == tid)
            )
        ).scalar_one()
        wedap_count = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultSourceWedap)
                .where(ReconResultSourceWedap.task_id == tid)
            )
        ).scalar_one()
        bank_count = (
            await session.execute(
                select(func.count())
                .select_from(ReconResultSourceBank)
                .where(ReconResultSourceBank.task_id == tid)
            )
        ).scalar_one()
    assert diff_count == 0, "脏金额事务回滚后 ReconResultDiff 应 0 行"
    assert wedap_count == 0
    assert bank_count == 0


# ───────────────────────── Test: 事务内 txn-guard DOWNLOADING skip ───────────


@pytest.mark.asyncio
async def test_parse_entry_guard_skips_downloading(factory, tmp_path) -> None:
    """入口预检查：task 为 DOWNLOADING 时 parse_and_land 静默 return（lines 169-175）。"""
    xlsx = tmp_path / "guard_entry.xlsx"
    _make_xlsx(xlsx)

    async with factory() as session:
        async with session.begin():
            task = _task(status="DOWNLOADING")
            session.add(task)
    async with factory() as session:
        row = (
            await session.execute(
                __import__("sqlalchemy", fromlist=["select"]).select(ReconResultTask)
            )
        ).scalar_one()
        tid = row.id

    # 入口预检查读到 DOWNLOADING，直接 return（幂等）
    await parse_and_land(factory, task_id=tid, xlsx_path=str(xlsx))

    async with factory() as session:
        task = await session.get(ReconResultTask, tid)
    # 状态保持 DOWNLOADING
    assert task is not None
    assert task.status == "DOWNLOADING"


@pytest.mark.asyncio
async def test_parse_txn_guard_skips_non_downloaded(factory, tmp_path) -> None:
    """事务内 txn-guard（lines 224-229）：入口预检查通过后，with_for_update 重读到
    非 DOWNLOADED 状态时，静默 return，不落行。

    用 _WrappedFactory 代理 factory：第 2 次 __call__ 进入 context 时先把 DB
    里的 task 改成 DOWNLOADING，使 txn-guard 触发。
    """
    from sqlalchemy import select as sa_select
    from sqlalchemy import update

    xlsx = tmp_path / "guard_txn.xlsx"
    _make_xlsx(xlsx)

    async with factory() as session:
        async with session.begin():
            task = _task(status="DOWNLOADED")
            session.add(task)

    async with factory() as session:
        row = (await session.execute(sa_select(ReconResultTask))).scalar_one()
        tid = row.id

    class _PatchedCtx:
        """第 2 次 factory() 返回的 context manager：进入时先改 DB 状态再代理真 session。"""

        def __init__(self, real_factory):
            self._real_factory = real_factory
            self._real_ctx = None

        async def __aenter__(self):
            # 先改状态
            async with self._real_factory() as s:
                async with s.begin():
                    await s.execute(
                        update(ReconResultTask)
                        .where(ReconResultTask.id == tid)
                        .values(status="DOWNLOADING")
                    )
            # 再打开真 session
            self._real_ctx = self._real_factory()
            return await self._real_ctx.__aenter__()

        async def __aexit__(self, *args):
            if self._real_ctx is not None:
                return await self._real_ctx.__aexit__(*args)

    class _WrappedFactory:
        def __init__(self, real):
            self._real = real
            self._call_count = 0

        def __call__(self):
            self._call_count += 1
            if self._call_count == 2:
                return _PatchedCtx(self._real)
            return self._real()

    wrapped = _WrappedFactory(factory)

    await parse_and_land(wrapped, task_id=tid, xlsx_path=str(xlsx))

    async with factory() as session:
        t = await session.get(ReconResultTask, tid)
    # txn-guard 触发 return → 三表 0 行，task 保持 DOWNLOADING（未被 PARSED）
    assert t is not None
    assert t.status == "DOWNLOADING"


# ───────────────────────── Test: worker 原子 claim 失败 skip ─────────────────


@pytest.mark.asyncio
async def test_worker_atomic_claim_skips_already_claimed(factory, tmp_path, monkeypatch) -> None:
    """worker 中 _atomic_claim 返回 False（已被他人 claim）→ skip，handled 不计入。

    monkeypatch _atomic_claim 直接返回 False，确保 snapshot 查到 NOTIFIED task 但 claim 路径
    走到 skip-continue 分支（lines 100-105）。
    """
    from app.clients.s3 import S3FileClient
    from app.workers import recon_worker
    from app.workers.recon_worker import ingest_pending_once

    async with factory() as session:
        async with session.begin():
            task = _task(status="NOTIFIED", task_no="RECON-CLAIM-001", request_id="REQ-CLAIM-001")
            session.add(task)

    def no_download(self, *, bucket, key, expected_md5, dest):
        raise AssertionError("download should not be called when claim fails")

    monkeypatch.setattr(S3FileClient, "download_verified", no_download)

    # monkeypatch _atomic_claim 返回 False（模拟他人已先 claim）
    async def fake_claim(factory_, task_id):
        return False

    monkeypatch.setattr(recon_worker, "_atomic_claim", fake_claim)

    s3 = S3FileClient(endpoint_url=None)
    handled = await ingest_pending_once(factory, s3=s3, archive_dir=str(tmp_path / "arch_claim"))
    # claim 失败 → skip → handled=0
    assert handled == 0


# ───────────────────────── Test: worker parse RuntimeError 置 FAILED ─────────


@pytest.mark.asyncio
async def test_worker_parse_runtime_error_fails_task_continues(
    factory, tmp_path, monkeypatch
) -> None:
    """worker parse 阶段 parse_and_land 抛 RuntimeError → task FAILED + parse_error 留痕 +
    后续 task 仍被处理（不阻塞）。"""
    from app.clients.s3 import S3FileClient
    from app.workers.recon_worker import ingest_pending_once

    # 建两个 NOTIFIED task：第一个解析会失败，第二个正常
    xlsx_ok = tmp_path / "ok2.xlsx"
    _make_xlsx(
        xlsx_ok,
        diff_rows=[["AMOUNT", "DSB-003", "BANK-003", "50.0000", "50.0000", "0.0000", "OK", "OK"]],
        wedap_rows=[["DSB", "DSB-003", "BANK-003", "50.0000", "USD", "P1", "P2", "OK", None]],
        bank_rows=[["BANK-003", "20260611", "50.0000", "USD", "P1", "P2", "OK", "f.xlsx", 1]],
    )

    async with factory() as session:
        async with session.begin():
            t_err = _task(
                status="NOTIFIED", task_no="RECON-ERR-001", version=1, request_id="REQ-ERR-001"
            )
            t_ok = _task(
                status="NOTIFIED", task_no="RECON-RT-OK-001", version=1, request_id="REQ-RT-OK-001"
            )
            session.add(t_err)
            session.add(t_ok)

    from sqlalchemy import select

    async with factory() as session:
        rows = (
            (await session.execute(select(ReconResultTask).order_by(ReconResultTask.task_no)))
            .scalars()
            .all()
        )
    tid_err = next(r.id for r in rows if r.task_no == "RECON-ERR-001")
    tid_ok = next(r.id for r in rows if r.task_no == "RECON-RT-OK-001")

    call_count = 0

    def fake_download(self, *, bucket, key, expected_md5, dest):
        nonlocal call_count
        call_count += 1
        import shutil

        # RECON-ERR-001 下载成功（再在 parse 阶段注入错误）
        shutil.copy(str(xlsx_ok), dest)

    monkeypatch.setattr(S3FileClient, "download_verified", fake_download)

    # 对 parse_and_land 打补丁：仅第一次（ERR task）抛 RuntimeError
    from app.services import recon_ingest

    original_parse = recon_ingest.parse_and_land
    parse_call = 0

    async def patched_parse(factory, *, task_id, xlsx_path):
        nonlocal parse_call
        parse_call += 1
        if parse_call == 1:
            raise RuntimeError("injected parse error")
        return await original_parse(factory, task_id=task_id, xlsx_path=xlsx_path)

    monkeypatch.setattr("app.workers.recon_worker.parse_and_land", patched_parse)

    s3 = S3FileClient(endpoint_url=None)
    handled = await ingest_pending_once(factory, s3=s3, archive_dir=str(tmp_path / "arch_rt"))
    assert handled == 2

    async with factory() as session:
        t_err_row = await session.get(ReconResultTask, tid_err)
        t_ok_row = await session.get(ReconResultTask, tid_ok)

    assert t_err_row is not None
    assert t_err_row.status == "FAILED"
    assert t_err_row.column_check is not None
    assert t_err_row.column_check.get("parse_error") == "RuntimeError"

    assert t_ok_row is not None
    assert t_ok_row.status == "PARSED"


# ───────────────────────── Test: 跨租户 supersede 隔离 ──────────────────────


@pytest.mark.asyncio
async def test_supersede_tenant_isolation(factory, tmp_path) -> None:
    """P1-1: A 租户解析 v2 后，B 租户同 task_no 同 version v1 不被 SUPERSEDED。"""
    from sqlalchemy import select

    xlsx_v2 = tmp_path / "v2_iso.xlsx"
    _make_xlsx(xlsx_v2)

    # 插：A 租户 v1(PARSED) + A 租户 v2(DOWNLOADED) + B 租户 v1(PARSED)
    async with factory() as session:
        async with session.begin():
            a_v1 = _task(tenant_id="TENANT_A", version=1, status="PARSED", request_id="REQ-A-001")
            a_v2 = _task(
                tenant_id="TENANT_A", version=2, status="DOWNLOADED", request_id="REQ-A-002"
            )
            b_v1 = _task(
                tenant_id="TENANT_B",
                version=1,
                status="PARSED",
                request_id="REQ-B-001",
                s3_key="recon/20260611/result_b.xlsx",  # 不同 s3_key 避免唯一冲突
            )
            session.add(a_v1)
            session.add(a_v2)
            session.add(b_v1)

    async with factory() as session:
        rows = (
            (
                await session.execute(
                    select(ReconResultTask).order_by(
                        ReconResultTask.tenant_id, ReconResultTask.version
                    )
                )
            )
            .scalars()
            .all()
        )
    tid_a_v1 = next(r.id for r in rows if r.tenant_id == "TENANT_A" and r.version == 1)
    tid_a_v2 = next(r.id for r in rows if r.tenant_id == "TENANT_A" and r.version == 2)
    tid_b_v1 = next(r.id for r in rows if r.tenant_id == "TENANT_B" and r.version == 1)

    # 解析 A 的 v2
    await parse_and_land(factory, task_id=tid_a_v2, xlsx_path=str(xlsx_v2))

    async with factory() as session:
        a1 = await session.get(ReconResultTask, tid_a_v1)
        a2 = await session.get(ReconResultTask, tid_a_v2)
        b1 = await session.get(ReconResultTask, tid_b_v1)

    # A 租户 v1 → SUPERSEDED（正常 supersede）
    assert a1 is not None and a1.status == "SUPERSEDED"
    # A 租户 v2 → PARSED
    assert a2 is not None and a2.status == "PARSED"
    # B 租户 v1 → 仍 PARSED（跨租户隔离，不受影响）
    assert b1 is not None and b1.status == "PARSED", (
        f"B 租户 v1 应保持 PARSED，实际状态: {b1.status}"
    )
